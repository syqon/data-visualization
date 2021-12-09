#!/usr/bin/env python
# coding: utf-8

# In[1]:


import pandas as pd
import seaborn as sns
import matplotlib.pyplot as plt
import numpy as np


# In[222]:


#读取2021数据
company = pd.read_csv("C:/Users/yinxi/Desktop/jupyter/1/1.csv")
#print(type(company))
company.head()


# In[163]:


#读取2020数据
company_2020 = pd.read_csv("C:/Users/yinxi/Desktop/jupyter/1/2020死亡公司.csv")
company_2020


# In[167]:


#读取2019数据
company_2019 = pd.read_csv("C:/Users/yinxi/Desktop/jupyter/1/2019死亡公司.csv")
company_2019


# In[7]:


pip install jieba


# In[6]:


#2021年死亡公司行业排名
#company = company[company['Industry'] == '广告营销']
#company


# In[7]:


from collections import Counter
from pprint import pprint

counter = Counter(company['Industry'])
# 打印2021前十行业的频次
pprint(counter.most_common(10))


# In[8]:


counter_2020 = Counter(company_2020['Industry'])
# 打印2020前十行业的频次
pprint(counter_2020.most_common(10))


# In[9]:


counter_2019 = Counter(company_2019['Industry'])
# 打印2019前十行业的频次
pprint(counter_2019.most_common(10))


# In[10]:


import json


# In[11]:


pip install -i https://pypi.tuna.tsinghua.edu.cn/simple pyecharts


# In[12]:


pip install -i https://pypi.tuna.tsinghua.edu.cn/simple pyecharts_snapshot


# In[13]:


get_ipython().system('pip install pyecharts')


# In[45]:


#2021前十行业柱状图
import matplotlib.pyplot as plt 
import numpy as np
X = ['E-Commerce', 'Education', 'Corporation service', 'Local life', 'Finance', 'Entertainment media','Car traffic','Social network','Game','Medical health']
Y = np.array([108,107,86,61,59,47,38,38,37,31])
Y_2020 = np.array([488,318,517,314,439,278,151,158,197,172])
Y_2019 = np.array([738,369,720,512,621,396,227,251,377,0])
plt.figure(figsize=(20,20), dpi=300)
plt.rc('font',family='Times New Roman') 
#plt.xticks(range(len(Y))
x = np.arange(len(Y))
width = 0.25
for i in range(len(x)):
    x[i]=x[i]+width
plt.bar(x-width,Y, width = width, label = '2021')
plt.bar(x, Y_2020, width=width, label='2020')
plt.bar(x+width, Y_2019, width=width, label='2019')
plt.yticks(fontsize=15)
plt.xticks(x, labels=X, fontsize = 15)
plt.xlabel('Industry',fontsize=18, weight = "bold")
plt.ylabel('Number',fontsize=18,weight = "bold")
plt.legend(borderpad=2, labelspacing=1, fontsize = 12)

plt.title("Number of death companies in Top 10 Industry during 2019-2021",fontsize=29, weight="bold", y = 1.05)

#显示数值 
for a,b in zip(x-width,Y): 
    plt.text(a,b,'%.0f'%b,ha='center',va='bottom',fontsize=12, weight = "bold");
for a,b in zip(x,Y_2020): 
    plt.text(a,b,'%.0f'%b,ha='center',va='bottom',fontsize=12, weight = "bold");
for a,b in zip(x+width,Y_2019): 
    plt.text(a,b,'%.0f'%b,ha='center',va='bottom',fontsize=12, weight = "bold");




plt.savefig("2019-2020-2021industry.jpg", dpi = 300)
plt.show()


# In[250]:


#2019,2020,2021前十行业柱状图
import matplotlib.pyplot as plt 

X = ['E-Commerce', 'Education', 'Corporation service', 'Local life', 'Finance', 'Entertainment media','Car traffic','Social network','Game','Medical health']
Y = np.array([108,107,86,61,59,47,38,38,37,31])
plt.figure(figsize=(15,15), dpi=300)
plt.rc('font',family='Times New Roman') 
#plt.xticks(range(len(Y))
plt.bar(X, Y,width = 0.5)
plt.legend(loc=[0,0])
plt.xticks(fontsize=10)
plt.xlabel('Industry',fontsize=18, weight = "bold")
plt.ylabel('Number',fontsize=18,weight = "bold")
plt.title("Top 10 Industry rankings of death companies in 2021",fontsize=29, weight="bold", y = 1.05)
#显示数值 
xx = np.arange(len(X));
for a,b in zip(xx,Y): 
     plt.text(a,b,'%.0f'%b,ha='center',va='bottom',fontsize=12, weight = "bold");


plt.savefig("2021industry.jpg", dpi = 300)
plt.show()


# In[49]:


get_ipython().system('pip install pyecharts')
get_ipython().system('pip install wheel')
get_ipython().system('pip install echarts-countries-pypkg')
get_ipython().system('pip install echarts-china-provinces-pypkg')
get_ipython().system('pip install echarts-china-cities-pypkg')
get_ipython().system('pip install echarts-china-counties-pypkg')
get_ipython().system('pip install echarts-china-misc-pypkg ')


# In[50]:


pip install -i https://pypi.tuna.tsinghua.edu.cn/simple pyecharts


# In[51]:


pip install -i https://pypi.tuna.tsinghua.edu.cn/simple pyecharts_snapshot


# In[52]:


#计算2021死亡公司前十地区
counter_place = Counter(company['Place'])
# 打印前十地区的频次
place = pprint(counter_place.most_common(10))


# In[53]:


#计算2020死亡公司前十地区
counter_place_2020 = Counter(company_2020['Place'])
# 打印前十地区的频次
place = pprint(counter_place_2020.most_common(10))


# In[54]:


#计算2019死亡公司前十地区
counter_place_2019 = Counter(company_2019['Place'])
# 打印前十地区的频次
place = pprint(counter_place_2019.most_common(10))


# In[55]:


#2019,2020, 2021前十死亡公司地区分布
from pyecharts import Map
#from pyecharts import province

province_distribution_2021 = {'北京': 204, '广东': 201, '上海': 116, '浙江': 57, '江苏': 51, '山东': 30, '湖北': 26, '重庆': 18, '天津': 15,'河南': 11}
province_distribution_2020 = {'北京': 1155, '广东': 859, '上海': 520, '浙江': 358, '江苏': 164, '湖北': 107, '四川': 105,'山东': 91, '福建': 83, '重庆': 55}
province_distribution_2019 = {'北京': 1836, '广东': 988, '上海': 757, '浙江': 425, '四川': 419, '江苏': 212, '湖北': 128, '陕西': 108, '福建': 99, '山东': 85}

province1 = list(province_distribution_2021.keys())
province2 = list(province_distribution_2020.keys())
province3 = list(province_distribution_2019.keys())
values1 = list(province_distribution_2021.values())
values2 = list(province_distribution_2020.values())
values3 = list(province_distribution_2019.values())

map = Map("Place distribution of death companies in 2019,2020,2021", 'China Map', width=1300, height=600)
map.add("2021", province1,values1, visual_range=[10, 2000], maptype='china', is_visualmap=True,
        visual_text_color='#000')
map.add("2020", province2,values2, visual_range=[10, 2000], maptype='china', is_visualmap=True,
        visual_text_color='#000')
map.add("2019", province3,values3, visual_range=[10, 2000], maptype='china', is_visualmap=True,
        visual_text_color='#000')
map.render(path="Place distribution of death companies.html")


# In[56]:


#2021前十死亡公司地区分布
from pyecharts import Map
#from pyecharts import province

province_distribution = {'北京': 204, '广东': 201, '上海': 116, '浙江': 57, '江苏': 51, '山东': 30, '湖北': 26, '重庆': 18, '天津': 15,'河南': 11}
province = list(province_distribution.keys())
values = list(province_distribution.values())

map = Map("Place distribution of death companies in 2021", 'China Map', width=1200, height=600)
map.add("China Map", province, values, visual_range=[0, 204], maptype='china', is_visualmap=True,
        visual_text_color='#000')
map.render(path="Place distribution of death companies in 2021.html")


# In[57]:


#2020死亡公司前十地区饼图
import matplotlib.pyplot as plt

#plt.rcParams['font.sans-serif']=['Microsoft YaHei']  #显示中文标签,处理中文乱码问题
#plt.rcParams['axes.unicode_minus']=False  #坐标轴负号的处理
plt.figure(figsize=(10,10), dpi=80)
plt.axes(aspect='equal')  #将横、纵坐标轴标准化处理，确保饼图是一个正圆
data = [1155, 869, 520, 358, 164,107,105,91,83,55]
labels = ['Beijing','Guangdong','Shanghai','Zhejiang','Jiangsu','Hubei','Sichuan','Shandong','Fujian','Chongqing']
explode = [0.2, 0.1, 0.05, 0.02,0,0,0,0,0,0]  #生成数据，用于凸显
colors = ['#9999ff', '#ff9999', '#7777aa', '#2442aa', '#dd5555'] 

plt.rc('font',family='Times New Roman') 
#mpl.rcParams['font.size']=16
plt.pie(x=data,  #绘图数据
        explode=explode, #指定饼图某些部分的突出显示，即呈现爆炸式
        labels=labels,  #添加教育水平标签
        colors=colors,
        autopct='%.2f%%',  #设置百分比的格式，这里保留两位小数
        pctdistance=0.8,  #设置百分比标签与圆心的距离
        labeldistance=1.1,  #设置水平标签与圆心的距离
        startangle=180,  #设置饼图的初始角度
        radius=3.5,  #设置饼图的半径
        counterclock=False,  #设置为顺时针方向
        wedgeprops={'linewidth':1, 'edgecolor':'green'},  #设置饼图内外边界的属性值
        textprops={'fontsize':12, 'color':'black'},  #设置文本标签的属性值
       )
plt.height = 50
plt.width = 50

# 显示图例
plt.legend(loc="center left",
           bbox_to_anchor=(3, -0.5, 0.3, 1), fontsize = 10)  
#图标题， 设置字号，位置
plt.title('Place distribution of death companies in 2020', fontsize = 16, weight="bold", y = 2.2)
plt.axis('off') 
plt.subplots_adjust(right =0.2)
plt.savefig("2020place.jpg", dpi = 300,bbox_inches = 'tight')
#显示图形
plt.show()


# In[58]:


#2021死亡公司前十地区饼图
import matplotlib.pyplot as plt

plt.figure(figsize=(10,10), dpi=200)
plt.axes(aspect='equal')  #将横、纵坐标轴标准化处理，确保饼图是一个正圆
data = [204, 201, 116, 57, 51,30,26,18,15,11]
labels = ['Beijing','Guangdong','Shanghai','Zhejiang','Jiangsu','Shandong','Hubei','Chongqing','Tianjin','Henan']
explode = [0.2, 0.1, 0.05, 0.02, 0,0,0,0,0,0]  #生成数据，用于凸显
colors = ['#9999ff', '#ff9999', '#7777aa', '#2442aa', '#dd5555'] 

plt.rc('font',family='Times New Roman') 
#mpl.rcParams['font.size']=16
plt.pie(x=data,  #绘图数据
        explode=explode, #指定饼图某些部分的突出显示，即呈现爆炸式
        labels=labels,  #添加教育水平标签
        colors=colors,
        autopct='%.2f%%',  #设置百分比的格式，这里保留两位小数
        pctdistance=0.8,  #设置百分比标签与圆心的距离
        labeldistance=1.1,  #设置水平标签与圆心的距离
        startangle=180,  #设置饼图的初始角度
        radius=3.5,  #设置饼图的半径
        counterclock=False,  #设置为顺时针方向
        wedgeprops={'linewidth':1, 'edgecolor':'green'},  #设置饼图内外边界的属性值
        textprops={'fontsize':12, 'color':'black'},  #设置文本标签的属性值
       )
plt.height = 50
plt.width = 50

# 显示图例
plt.legend(loc="center left",
           bbox_to_anchor=(3, -0.5, 0.3, 1), fontsize = 10)  
#图标题， 设置字号，位置
plt.title('Place distribution of death companies in 2021', fontsize = 16, weight="bold", y = 2.2)
plt.axis('off') 
plt.subplots_adjust(right =0.2)
plt.savefig("2021place.jpg", dpi = 300,bbox_inches = 'tight')
#显示图形
plt.show()


# In[59]:


#2019死亡公司前十地区饼图
import matplotlib.pyplot as plt

plt.figure(figsize=(10,10), dpi=80)
plt.axes(aspect='equal')  #将横、纵坐标轴标准化处理，确保饼图是一个正圆
data = [1836, 988, 757, 425, 419,212,128,108,99,85]
labels = ['Beijing','Guangdong','Shanghai','Zhejiang','Sichuan','Jiangsu','Hubei','Shanxi','Fujian','Shandong']
explode = [0.2, 0.1, 0.05, 0.02,0,0,0,0,0,0]  #生成数据，用于凸显
colors = ['#9999ff', '#ff9999', '#7777aa', '#2442aa', '#dd5555'] 

plt.rc('font',family='Times New Roman') 
#mpl.rcParams['font.size']=16
plt.pie(x=data,  #绘图数据
        explode=explode, #指定饼图某些部分的突出显示，即呈现爆炸式
        labels=labels,  #添加教育水平标签
        colors=colors,
        autopct='%.2f%%',  #设置百分比的格式，这里保留两位小数
        pctdistance=0.8,  #设置百分比标签与圆心的距离
        labeldistance=1.1,  #设置水平标签与圆心的距离
        startangle=180,  #设置饼图的初始角度
        radius=3.5,  #设置饼图的半径
        counterclock=False,  #设置为顺时针方向
        wedgeprops={'linewidth':1, 'edgecolor':'green'},  #设置饼图内外边界的属性值
        textprops={'fontsize':12, 'color':'black'},  #设置文本标签的属性值
       )
plt.height = 50
plt.width = 50

# 显示图例
plt.legend(loc="center left",
           bbox_to_anchor=(3, -0.5, 0.3, 1), fontsize = 10)  
#图标题， 设置字号，位置
plt.title('Place distribution of death companies in 2019', fontsize = 16, weight="bold", y = 2.2)
plt.axis('off') 
plt.subplots_adjust(right =0.2)
plt.savefig("2019place.jpg", dpi = 300,bbox_inches = 'tight')
#显示图形
plt.show()


# In[60]:


import numpy as np
import pandas as pd
from pandas import Series, DataFrame
get_ipython().run_line_magic('matplotlib', 'inline')
import matplotlib.pyplot as plt
from pylab import mpl


# In[251]:


#2019，2020和2021死亡公司数量对比柱状图
import matplotlib.pyplot as plt 

X = ['2019', '2020', '2021']
Y = np.array([5646,3922,817])
plt.figure(figsize=(10,10), dpi=300)
plt.rc('font',family='Times New Roman') 
#plt.xticks(range(len(Y))
plt.bar(X, Y,width = 0.3, color = ['orange', 'blue', 'green'] )
plt.legend(loc=[0,0], fontsize = 12)
plt.xticks(fontsize=14)
plt.xlabel('Year',fontsize=18, weight = "bold")
plt.ylabel('Number',fontsize=18,weight = "bold")
plt.title("Comparision of number of death companies in 2019-2021",fontsize=20, weight="bold", y = 1.05)
#显示数值 
xx = np.arange(len(X));
for a,b in zip(xx,Y): 
     plt.text(a,b,'%.0f'%b,ha='center',va='bottom',fontsize=10, weight = "bold");


plt.savefig("number comparison.jpg", dpi = 300)
plt.show()


# In[ ]:


#2021前十行业柱状图
import matplotlib.pyplot as plt 
import numpy as np
X = ['E-Commerce', 'Education', 'Corporation service', 'Local life', 'Finance', 'Entertainment media','Car traffic','Social network','Game','Medical health']
Y = np.array([108,107,86,61,59,47,38,38,37,31])
Y_2020 = np.array([488,318,517,314,439,278,151,158,197,172])
Y_2019 = np.array([738,369,720,512,621,396,227,251,377,0])
plt.figure(figsize=(20,20), dpi=300)
plt.rc('font',family='Times New Roman') 
#plt.xticks(range(len(Y))
x = np.arange(len(Y))
width = 0.25
for i in range(len(x)):
    x[i]=x[i]+width
plt.bar(x-width,Y, width = width, label = '2021')
plt.bar(x, Y_2020, width=width, label='2020')
plt.bar(x+width, Y_2019, width=width, label='2019')
plt.yticks(fontsize=15)
plt.xticks(x, labels=labels, fontsize = 15)
plt.ylabel('Number',fontsize=15)
plt.legend(loc = [0,0])

plt.title("Top 10 Industry rankings of death companies in 2019,2020,2021",fontsize=29, weight="bold", y = 1.05)
#显示数值 
xx = np.arange(len(X));
for a,b in zip(xx,Y): 
     plt.text(a,b,'%.0f'%b,ha='center',va='bottom',fontsize=12, weight = "bold");
for a1,b1 in zip(xx,Y_2020): 
    plt.text(a1,b1,'%.0f'%b,ha='center',va='bottom',fontsize=12, weight = "bold");


plt.savefig("2019-2020-2021industry.jpg", dpi = 300)
plt.show()


# In[64]:


#提取2021列表中Close time列
close_time = company['Close time']
close_time


# In[65]:


import re


# In[66]:


#计算2021每月死亡公司数量频次
from collections import Counter
from pprint import pprint

close_time_month_2021 = Counter(company['Close time month'])
# 打印每月的频次
pprint(close_time_month_2021)


# In[67]:


#计算2020每月死亡公司数量频次
close_time_month_2020 = Counter(company_2020['Close time month'])
# 打印每月的频次
pprint(close_time_month_2020)


# In[68]:


#计算2019每月死亡公司数量频次
close_time_month_2019 = Counter(company_2019['Close time month'])
# 打印每月的频次
pprint(close_time_month_2019)


# In[ ]:


#提取年+月份
#print(type(close_time))
#close_time_month = company['Close time'].str.extract(r'^(\S{7})', expand= False)
#close_time_month = company['Close time'].str.extract(r'^(\S{6})+$(\S{7})', expand= False) 
#print(close_time_month)


# In[280]:


#带月份趋势变换的2019-2021死亡公司数量折线图
import seaborn as sns
import matplotlib.pyplot as plt
import pandas as pd
import numpy as np

data=pd.read_csv('C:/Users/yinxi/Desktop/jupyter/1/month.csv')
xdata=[]
y1data=[]
y2data=[]
y3data=[]
xdata=data.loc[:,'Month']
y1data=data.loc[:,'Number_2019']
y2data=data.loc[:,'Number_2020']
y3data=data.loc[:,'Number_2021']
plt.plot(xdata,y1data,'o-',label='2019',linewidth=1)
plt.plot(xdata,y2data,'--',label='2020',linewidth=1)
plt.plot(xdata,y3data,'-',label='2021',linewidth=1)
plt.legend
xx1 = np.arange(len(xdata));#显示数值
for a,b in zip(xx1,y1data): 
     plt.text(a,b,'%.0f'%b,ha='center',va='bottom',fontsize=8);
xx2 = np.arange(len(xdata));#显示数值
for a,b in zip(xx2,y2data): 
     plt.text(a,b,'%.0f'%b,ha='center',va='bottom',fontsize=8);
xx3 = np.arange(len(xdata));#显示数值
for a,b in zip(xx3,y3data): 
     plt.text(a,b,'%.0f'%b,ha='center',va='bottom',fontsize=8);
plt.xlabel(u"Month", size = 12)
plt.ylabel(u"Number",size = 12)
plt.grid(True, linestyle="--", alpha=10)
plt.legend(loc="best", fontsize = 14)
plt.title(u"Trend of Number of Death Companies during 2019-2021",size=15)
plt.savefig("month comparision.jpg", dpi = 300)#保存
plt.figure(figsize=(50, 25), dpi=300)
plt.show()


# In[142]:


#2021读取数据
company = pd.read_csv("C:/Users/yinxi/Desktop/jupyter/1/1.csv")
#print(type(company))
company.head()


# In[150]:


#2021死亡原因词云图
result=[]
list1 = (company['Death reason'])
list2 = (company['Death reason2'])
list3 = (company['Death reason3'])
list4 = (company['Death reason4'])
list5 = (company['Death reason5'])
list6 = (company['Death reason6'])
for element in list1:
    result.append(element)
for element in list2:
    result.append(element)
for element in list3:
    result.append(element)
for element in list4:
    result.append(element)
for element in list5:
    result.append(element)
for element in list6:
    result.append(element)
print(result)


# In[72]:


#data cleaning：去掉nan
new_result = [x for x in result if pd.isnull(x) == False]
print(new_result)


# In[73]:


from collections import Counter
from pprint import pprint

counter = Counter(new_result)

# 打印2021死亡原因前二十高频词
Frequency=pprint(counter.most_common(20))
print=(counter)


# In[74]:


get_ipython().system('pip install wordcloud')


# In[7]:


import pandas as pd
#death reason读取数据
death_reason = pd.read_csv("C:/Users/yinxi/Desktop/jupyter/1/death reason.csv")
#print(type(company))
death_reason.head()


# In[8]:


#转成字典
from openpyxl import load_workbook
look_up_table_path='C:/Users/yinxi/Desktop/jupyter/1/death reason.xlsx'
look_up_table_row_start=2
look_up_table_row_number=17


# In[9]:


reason_frequency_dict={}
look_up_table_excel=load_workbook(look_up_table_path)
look_up_table_all_sheet=look_up_table_excel.get_sheet_names()
look_up_table_sheet=look_up_table_excel.get_sheet_by_name(look_up_table_all_sheet[0])

for i in range(look_up_table_row_start,look_up_table_row_start+look_up_table_row_number-1):
    reason=look_up_table_sheet.cell(i,1).value
    frequency=look_up_table_sheet.cell(i,2).value
    reason_frequency_dict[reason]=frequency


# In[10]:


reason_frequency_dict


# In[15]:


from matplotlib import pyplot as plt
import random
from pylab import mpl
import numpy as np
from matplotlib.pyplot import MultipleLocator
from wordcloud import WordCloud,ImageColorGenerator
from matplotlib import colors
color_list=['#CD853F','#DC143C','#00FF7F','#FF6347','#8B008B','#00FFFF','#0000FF','#8B0000','#FF8C00',
            '#1E90FF','#00FF00','#FFD700','#008080','#008B8B','#8A2BE2','#228B22','#FA8072','#808080']
#调用
colormap=colors.ListedColormap(color_list)

plt.figure(figsize=(20, 15), dpi=200)
wordcloud = WordCloud(font_path='SimHei.ttf', background_color="white",width =4000,height= 2000,margin= 10, colormap=colormap, random_state=18).fit_words(reason_frequency_dict)
plt.imshow(wordcloud)
plt.axis("off")
plt.title("Top 20 death reason of death companies during 2019-2021", weight = "bold", fontsize=30, y = 1.1)
plt.savefig("death reason.jpg", dpi = 300)
# 显示
plt.show()


# In[8]:


from collections import Counter
from pprint import pprint

counter = Counter(company['Life_time_year'])
# 打印2021存活时间频次
pprint(counter)


# In[16]:


from collections import Counter
from pprint import pprint

counter_2020 = Counter(company_2020['Life_time_year'])
# 打印2020存活时间频次
pprint(counter_2020)


# In[13]:


from collections import Counter
from pprint import pprint

counter_2019 = Counter(company_2019['Life_time_year'])
# 打印2019存活时间频次
pprint(counter_2019)


# In[25]:


#2019-2021存活时间分布饼图
import matplotlib.pyplot as plt

plt.figure(figsize=(12,12), dpi=80)
plt.axes(aspect='equal')  #将横、纵坐标轴标准化处理，确保饼图是一个正圆
data = [263, 4073, 4574, 1055, 222,198]
labels = ['<1 year','1-3 years','4-6 years','7-9 years','9-12 years','>12 years']
explode = [0.2, 0.155, 0.15, 0.1,0.05,0]  #生成数据，用于凸显
colors = ['#9999ff', '#ff9999', '#7777aa', '#2442aa', '#dd5555','#FF9933'] 

plt.rc('font',family='Times New Roman') 
#mpl.rcParams['font.size']=16
plt.pie(x=data,  #绘图数据
        explode=explode, #指定饼图某些部分的突出显示，即呈现爆炸式
        labels=labels,  #添加教育水平标签
        colors=colors,
        autopct='%.2f%%',  #设置百分比的格式，这里保留两位小数
        pctdistance=0.8,  #设置百分比标签与圆心的距离
        labeldistance=1.1,  #设置水平标签与圆心的距离
        startangle=180,  #设置饼图的初始角度
        radius=3.5,  #设置饼图的半径
        counterclock=False,  #设置为顺时针方向
        wedgeprops={'linewidth':1, 'edgecolor':'green'},  #设置饼图内外边界的属性值
        textprops={'fontsize':12, 'color':'black'},  #设置文本标签的属性值
       )
plt.height = 50
plt.width = 50

# 显示图例
plt.legend(loc="center left",
           bbox_to_anchor=(3, -0.5, 0.3, 1), fontsize = 10)  
#图标题， 设置字号，位置
plt.title('Life time distribution of death companies in 2019-2021', fontsize = 16, weight="bold", y = 2.2)
plt.axis('off') 
plt.subplots_adjust(right =0.2)
plt.savefig("2019-2021lifetime.jpg", dpi = 300,bbox_inches = 'tight')
#显示图形
plt.show()


# In[26]:


counter= Counter(company['Cast state'])
# 打印2021获投情况
pprint(counter)


# In[27]:


counter_2020 = Counter(company_2020['Cast state'])
# 打印2020获投情况
pprint(counter_2020)


# In[28]:


counter_2019 = Counter(company_2019['Cast state'])
# 打印2019获投情况
pprint(counter_2019)


# In[262]:


#2019-2021获投情况有/无分布饼图
import matplotlib.pyplot as plt

plt.figure(figsize=(12,12), dpi=80)
plt.axes(aspect='equal')  #将横、纵坐标轴标准化处理，确保饼图是一个正圆
data = [5911, 4470]
labels = ['No financing','Financing']
explode = [0.1, 0]  #生成数据，用于凸显
colors = ['#9999ff', '#ff9999'] 

plt.rc('font',family='Times New Roman') 
#mpl.rcParams['font.size']=16
plt.pie(x=data,  #绘图数据
        explode=explode, #指定饼图某些部分的突出显示，即呈现爆炸式
        labels=labels,  #添加教育水平标签
        colors=colors,
        autopct='%.2f%%',  #设置百分比的格式，这里保留两位小数
        pctdistance=0.5,  #设置百分比标签与圆心的距离
        labeldistance=1.1,  #设置水平标签与圆心的距离
        startangle=90,  #设置饼图的初始角度
        radius=3.5,  #设置饼图的半径
        counterclock=False,  #设置为顺时针方向
        wedgeprops={'linewidth':1, 'edgecolor':'green'},  #设置饼图内外边界的属性值
        textprops={'fontsize':12, 'color':'black'},  #设置文本标签的属性值
       )
plt.height = 50
plt.width = 50

# 显示图例
plt.legend(loc="center left",
           bbox_to_anchor=(4, -0.5, 0.3, 1), fontsize = 10)  
#图标题， 设置字号，位置
plt.title('Financing status of death companies in 2019-2021', fontsize = 16, weight="bold", y = 2.2)
plt.axis('off') 
plt.subplots_adjust(right =0.2)
plt.savefig("2019-2021financing status-1.jpg", dpi = 300,bbox_inches = 'tight')
#显示图形
plt.show()


# In[264]:


#2019-2021获投情况饼图
import matplotlib.pyplot as plt

plt.figure(figsize=(60,40), dpi=200)
plt.axes(aspect='equal')  #将横、纵坐标轴标准化处理，确保饼图是一个正圆
data = [1501, 199, 832, 97, 23,15, 88, 101, 1614]
labels = ['Angel round investment','Seed round','A Round/Pre-A Round/A+ Round',
          'B Round/B+ Round','C Round/C+ Round','D Round/D+ Round/E Round', 'Strategic investment', 'Unclear','Others (NEEQ; Listed; Amalgamated; Acquired; Delisted)']
#explode = [0, 0, 0.15, 0.1,0.05,0.05,0.1,0.1,0.1]  #生成数据，用于凸显
colors = ['#9999ff', '#ff9999', '#7777aa', '#2442aa', '#dd5555','#FF9933', '#6699CC', '#339999', '#FF99CC'] 

plt.rc('font',family='Times New Roman') 
#mpl.rcParams['font.size']=16
plt.pie(x=data,  #绘图数据
        #explode=explode, #指定饼图某些部分的突出显示，即呈现爆炸式
        labels=labels,  #添加教育水平标签
        colors=colors,
        autopct='%.2f%%',  #设置百分比的格式，这里保留两位小数
        pctdistance=0.8,  #设置百分比标签与圆心的距离
        labeldistance=1.05,  #设置水平标签与圆心的距离
        startangle=90,  #设置饼图的初始角度
        radius=4,  #设置饼图的半径
        counterclock=False,  #设置为顺时针方向
        wedgeprops={'linewidth':1, 'edgecolor':'green'},  #设置饼图内外边界的属性值
        textprops={'fontsize':13, 'color':'black', 'weight':'bold'},  #设置文本标签的属性值
       )
plt.height = 50
plt.width = 50


    # 显示图例
plt.legend(loc="center left",
           bbox_to_anchor=(3, -0.5, 0.3, 1), fontsize = 28)  
#图标题， 设置字号，位置
plt.title('Financing status distribution of death companies in 2019-2021', fontsize = 40, weight="bold", y = 2.2)
plt.axis('off') 
plt.subplots_adjust(right =0.2)
plt.savefig("2019-2021Financing status-2.jpg", dpi = 300, bbox_inches = 'tight')
#显示图形
plt.show()


# In[100]:


#2021年未获投公司
company_1 = company[company['Cast state'] == '尚未获投']
#print(company_1)

#2021年未获投公司存活天数
company_2 = company_1['Life_time_day']
company_2 = list(company_2)
print(company_2)


# In[96]:


#2021年天使轮公司
company_1_1 = company[company['Cast state'] == '天使轮']
#print(company_1)

#2021年天使轮公司存活天数
company_2_1 = company_1_1['Life_time_day']
company_2_1 = list(company_2_1)
print(company_2_1)


# In[95]:


#2021年A轮公司
company_1_2 = company[company['Cast state'] == 'A轮']
#print(company_1)
#2021年A轮公司存活天数
company_2_2 = company_1_2['Life_time_day']
company_2_2 = list(company_2_2)
print(company_2_2)


# In[94]:


#2021年B轮公司
company_1_4 = company[company['Cast state'] == 'B轮']
#print(company_1)
#2021年B轮公司存活天数
company_2_4 = company_1_4['Life_time_day']
company_2_4 = list(company_2_4)
print(company_2_4)


# In[93]:


#2021年种子轮公司
company_1_3 = company[company['Cast state'] == '种子轮']
#print(company_1)
#2021年种子轮公司存活天数
company_2_3 = company_1_3['Life_time_day']
company_2_3 = list(company_2_3)
print(company_2_3)


# In[92]:


#2021年C轮公司
company_1_5 = company[company['Cast state'] == 'C轮']
#print(company_1)
#2021年C公司存活天数
company_2_5 = company_1_5['Life_time_day']
company_2_5 = list(company_2_5)
print(company_2_5)


# In[91]:


#2021年D轮公司
company_1_6 = company[company['Cast state'] == 'D轮']
#print(company_1)
#2021年D轮公司存活天数
company_2_6 = company_1_6['Life_time_day']
company_2_6 = list(company_2_6)
print(company_2_6)


# In[89]:


#2021年战略投资公司
company_1_7 = company[company['Cast state'] == '战略投资']
#print(company_1)
#2021年战略投资公司存活天数
company_2_7 = company_1_7['Life_time_day']
company_2_7 = list(company_2_7)
print(company_2_7)


# In[266]:


#2021融资情况---存活时间箱型图
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from pylab import *

plt.figure(figsize=(13,13), dpi=200)
rcParams['axes.unicode_minus'] = False
rcParams['font.sans-serif'] = ['Times New Roman']

y1 = pd.Series(np.array(company_2))
y2 = pd.Series(np.array(company_2_1))
y3 = pd.Series(np.array(company_2_3))
y4  = pd.Series(np.array(company_2_2))
y5  = pd.Series(np.array(company_2_4))
y6  = pd.Series(np.array(company_2_5))
y7  = pd.Series(np.array(company_2_6))
y8 = pd.Series(np.array(company_2_7))

data = pd.DataFrame({"No Financing": y1,
                     "Angel Round investment ": y2, 
                     "Seed investment": y3, 
                     "A Round": y4,
                     "B Round": y5,
                     "C Round": y6,
                     "D Round": y7,
                     "Strategic investment": y8,
                    })
#x = data['Uncasted'], data['Angel Round investment'], data['Seed investment'], data['A Round'], data['B Round'], data['C Round'], data['D Round'], data['Strategic investment'], 
#box_1, box_2, box_3, box_4, box_5, box_6, box_7, box_8 = data['Uncasted'], data['Angel Round investment'], data['Seed investment'], data['A Round'], data['B Round'], data['C Round'], data['D Round'], data['Strategic investment']

data.boxplot()

plt.height = 30
plt.width = 30

#图标题， 设置字号，位置
plt.title('Financing status and Life time (day) of death companies in 2021', fontsize = 20, weight="bold", y = 1.01)
plt.ylabel("Life time (day)", fontsize = 14, weight = "bold")
plt.xlabel("Financing status", fontsize = 14, weight = "bold")
plt.savefig("2021Financing status-lifetime.jpg", dpi = 300, bbox_inches = 'tight')
plt.show()


# In[168]:


#2020年未获投公司
company_1 = company_2020[company_2020['Cast state'] == '尚未获投']
#print(company_1)

#2020年未获投公司存活天数
company_2 = company_1['Life_time_day']
company_2 = list(company_2)
print(company_2)


# In[169]:


#2020年天使轮公司
company_1_1 = company_2020[company_2020['Cast state'] == '天使轮']
#print(company_1)

#2020年天使轮公司存活天数
company_2_1 = company_1_1['Life_time_day']
company_2_1 = list(company_2_1)
print(company_2_1)


# In[170]:


#2020年A轮公司
company_1_2 = company_2020[company_2020['Cast state'] == 'A轮']
#print(company_1)
#2020年A轮公司存活天数
company_2_2 = company_1_2['Life_time_day']
company_2_2 = list(company_2_2)
print(company_2_2)


# In[186]:


#2020年B轮公司
company_1_4 = company_2020[company_2020['Cast state'] == 'B轮']
#print(company_1)
#2020年B轮公司存活天数
company_2_4 = company_1_4['Life_time_day']
company_2_4 = list(company_2_4)
print(company_2_4)


# In[187]:


#2020年C轮公司
company_1_5 = company_2020[company_2020['Cast state'] == 'C轮']
#print(company_1)
#2020年C轮公司存活天数
company_2_5 = company_1_5['Life_time_day']
company_2_5 = list(company_2_5)
print(company_2_5)


# In[188]:


#2020年D轮公司
company_1_6 = company_2020[company_2020['Cast state'] == 'D轮']
#print(company_1)
#2020年D轮公司存活天数
company_2_6 = company_1_6['Life_time_day']
company_2_6 = list(company_2_6)
print(company_2_6)


# In[189]:


#2020年战略投资公司
company_1_7 = company_2020[company_2020['Cast state'] == '战略投资']
#print(company_1)
#2020年战略投资公司存活天数
company_2_7 = company_1_7['Life_time_day']
company_2_7 = list(company_2_7)
print(company_2_7)


# In[268]:


#2020融资情况---存活时间箱型图
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from pylab import *

plt.figure(figsize=(13,13), dpi=200)
rcParams['axes.unicode_minus'] = False
rcParams['font.sans-serif'] = ['Times New Roman']

y1 = pd.Series(np.array(company_2))
y2 = pd.Series(np.array(company_2_1))
y3 = pd.Series(np.array(company_2_3))
y4  = pd.Series(np.array(company_2_2))
y5  = pd.Series(np.array(company_2_4))
y6  = pd.Series(np.array(company_2_5))
y7  = pd.Series(np.array(company_2_6))
y8 = pd.Series(np.array(company_2_7))

data = pd.DataFrame({"No Financing": y1,
                     "Angel Round investment ": y2, 
                     "Seed investment": y3, 
                     "A Round": y4,
                     "B Round": y5,
                     "C Round": y6,
                     "D Round": y7,
                     "Strategic investment": y8,
                    })
#x = data['Uncasted'], data['Angel Round investment'], data['Seed investment'], data['A Round'], data['B Round'], data['C Round'], data['D Round'], data['Strategic investment'], 
#box_1, box_2, box_3, box_4, box_5, box_6, box_7, box_8 = data['Uncasted'], data['Angel Round investment'], data['Seed investment'], data['A Round'], data['B Round'], data['C Round'], data['D Round'], data['Strategic investment']

data.boxplot()

plt.height = 30
plt.width = 30

#图标题， 设置字号，位置
plt.title('Financing status and Life time (day) of death companies in 2020', fontsize = 20, weight="bold", y = 1.01)
plt.ylabel("Life time (day)", fontsize = 14, weight = "bold")
plt.xlabel("Financing status", fontsize = 14, weight = "bold")
plt.savefig("2020Financing status-lifetime.jpg", dpi = 300, bbox_inches = 'tight')
plt.show()


# In[271]:


#2019年未获投公司
company_1 = company_2019[company_2019['Cast state'] == '尚未获投']
#print(company_1)

#2021年未获投公司存活天数
company_2 = company_1['Life_time_day']
company_2 = list(company_2)
print(company_2)


# In[272]:


#2019年天使轮公司
company_1_1 = company_2019[company_2019['Cast state'] == '天使轮']
#print(company_1)

#2019年天使轮公司存活天数
company_2_1 = company_1_1['Life_time_day']
company_2_1 = list(company_2_1)
print(company_2_1)


# In[273]:


#2019年种子轮公司
company_1_3 = company_2019[company_2019['Cast state'] == '种子轮']
#print(company_1)
#2019年种子轮公司存活天数
company_2_3 = company_1_3['Life_time_day']
company_2_3 = list(company_2_3)
print(company_2_3)


# In[274]:


#2019年A轮公司
company_1_2 = company_2019[company_2019['Cast state'] == 'A轮']
#print(company_1)
#2019年A轮公司存活天数
company_2_2 = company_1_2['Life_time_day']
company_2_2 = list(company_2_2)
print(company_2_2)


# In[275]:


#2019年B轮公司
company_1_4 = company_2019[company_2019['Cast state'] == 'B轮']
#print(company_1)
#2019年B轮公司存活天数
company_2_4 = company_1_4['Life_time_day']
company_2_4 = list(company_2_4)
print(company_2_4)


# In[276]:


#2019年C轮公司
company_1_5 = company_2019[company_2019['Cast state'] == 'C轮']
#print(company_1)
#2019年C轮公司存活天数
company_2_5 = company_1_5['Life_time_day']
company_2_5 = list(company_2_5)
print(company_2_5)


# In[277]:


#2019年D轮公司
company_1_6 = company_2019[company_2019['Cast state'] == 'D轮']
#print(company_1)
#2019年D轮公司存活天数
company_2_6 = company_1_6['Life_time_day']
company_2_6 = list(company_2_6)
print(company_2_6)


# In[278]:


#2019年战略投资公司
company_1_7 = company_2019[company_2019['Cast state'] == '战略投资']
#print(company_1)
#2021年战略投资公司存活天数
company_2_7 = company_1_7['Life_time_day']
company_2_7 = list(company_2_7)
print(company_2_7)


# In[279]:


#2019融资情况---存活时间箱型图
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from pylab import *

plt.figure(figsize=(13,13), dpi=200)
rcParams['axes.unicode_minus'] = False
rcParams['font.sans-serif'] = ['Times New Roman']

y1 = pd.Series(np.array(company_2))
y2 = pd.Series(np.array(company_2_1))
y3 = pd.Series(np.array(company_2_3))
y4  = pd.Series(np.array(company_2_2))
y5  = pd.Series(np.array(company_2_4))
y6  = pd.Series(np.array(company_2_5))
y7  = pd.Series(np.array(company_2_6))
y8 = pd.Series(np.array(company_2_7))

data = pd.DataFrame({"No Financing": y1,
                     "Angel Round investment ": y2, 
                     "Seed investment": y3, 
                     "A Round": y4,
                     "B Round": y5,
                     "C Round": y6,
                     "D Round": y7,
                     "Strategic investment": y8,
                    })
#x = data['Uncasted'], data['Angel Round investment'], data['Seed investment'], data['A Round'], data['B Round'], data['C Round'], data['D Round'], data['Strategic investment'], 
#box_1, box_2, box_3, box_4, box_5, box_6, box_7, box_8 = data['Uncasted'], data['Angel Round investment'], data['Seed investment'], data['A Round'], data['B Round'], data['C Round'], data['D Round'], data['Strategic investment']

data.boxplot()

plt.height = 30
plt.width = 30

#图标题， 设置字号，位置
plt.title('Financing status and Life time (day) of death companies in 2019', fontsize = 20, weight="bold", y = 1.01)
plt.ylabel("Life time (day)", fontsize = 14, weight = "bold")
plt.xlabel("Financing status", fontsize = 14, weight = "bold")
plt.savefig("2019Financing status-lifetime.jpg", dpi = 300, bbox_inches = 'tight')
plt.show()


# In[202]:


#计算2021死亡公司前十行业
counter_industry = Counter(company['Industry'])
# 打印前十行业的频次
industry = pprint(counter_industry.most_common(10))


# In[208]:


#2021企业服务行业存活时间
industry_1 = company[company['Industry'] == '企业服务']
#print(industry_1)

industry_1_1 = industry_1['Life_time_day']
industry_1_1 = list(industry_1_1)
print(industry_1_1)


# In[209]:


#2021电子商务行业存活时间
industry_2 = company[company['Industry'] == '电子商务']

industry_2_2 = industry_2['Life_time_day']
industry_2_2 = list(industry_2_2)
print(industry_2_2)


# In[210]:


#2021教育行业存活时间
industry_3 = company[company['Industry'] == '教育']

industry_3_3 = industry_3['Life_time_day']
industry_3_3 = list(industry_3_3)
print(industry_3_3)


# In[211]:


#2021本地生活行业存活时间
industry_4 = company[company['Industry'] == '本地生活']

industry_4_4 = industry_4['Life_time_day']
industry_4_4 = list(industry_4_4)
print(industry_4_4)


# In[212]:


#2021金融行业存活时间
industry_5 = company[company['Industry'] == '金融']

industry_5_5 = industry_5['Life_time_day']
industry_5_5 = list(industry_5_5)
print(industry_5_5)


# In[217]:


#2021文娱传媒行业存活时间
industry_6 = company[company['Industry'] == '文娱传媒']

industry_6_6 = industry_6['Life_time_day']
industry_6_6 = list(industry_6_6)
print(industry_6_6)


# In[218]:


#2021社交网络行业存活时间
industry_7 = company[company['Industry'] == '社交网络']

industry_7_7 = industry_7['Life_time_day']
industry_7_7 = list(industry_7_7)
print(industry_7_7)


# In[219]:


#2021游戏行业存活时间
industry_8 = company[company['Industry'] == '游戏']

industry_8_8 = industry_8['Life_time_day']
industry_8_8 = list(industry_8_8)
print(industry_8_8)


# In[220]:


#2021汽车交通行业存活时间
industry_9 = company[company['Industry'] == '汽车交通']

industry_9_9 = industry_9['Life_time_day']
industry_9_9 = list(industry_9_9)
print(industry_9_9)


# In[223]:


#2021医疗健康行业存活时间
industry_10 = company[company['Industry'] == '医疗健康']

industry_10_10 = industry_10['Life_time_day']
industry_10_10 = list(industry_10_10)
print(industry_10_10)


# In[224]:


#2020企业服务行业存活时间
industry_1 = company_2020[company_2020['Industry'] == '企业服务']
#print(industry_1)

industry_1_2 = industry_1['Life_time_day']
industry_1_2 = list(industry_1_2)
print(industry_1_2)


# In[226]:


#2020电子商务行业存活时间
industry_2 = company_2020[company_2020['Industry'] == '电子商务']

industry_2_3 = industry_2['Life_time_day']
industry_2_3 = list(industry_2_3)
print(industry_2_3)


# In[227]:


#2020教育行业存活时间
industry_3 = company_2020[company_2020['Industry'] == '教育']

industry_3_4 = industry_3['Life_time_day']
industry_3_4 = list(industry_3_4)
print(industry_3_4)


# In[228]:


#2020本地生活行业存活时间
industry_4 = company_2020[company_2020['Industry'] == '本地生活']

industry_4_5 = industry_4['Life_time_day']
industry_4_5 = list(industry_4_5)
print(industry_4_5)


# In[229]:


#2020金融行业存活时间
industry_5 = company_2020[company_2020['Industry'] == '金融']

industry_5_6 = industry_5['Life_time_day']
industry_5_6 = list(industry_5_6)
print(industry_5_6)


# In[230]:


#2020文娱传媒行业存活时间
industry_6 = company_2020[company_2020['Industry'] == '文娱传媒']

industry_6_7 = industry_6['Life_time_day']
industry_6_7 = list(industry_6_7)
print(industry_6_7)


# In[231]:


#2020社交网络行业存活时间
industry_7 = company_2020[company_2020['Industry'] == '社交网络']

industry_7_8 = industry_7['Life_time_day']
industry_7_8 = list(industry_7_8)
print(industry_7_8)


# In[232]:


#2020游戏行业存活时间
industry_8 = company_2020[company_2020['Industry'] == '游戏']

industry_8_9 = industry_8['Life_time_day']
industry_8_9 = list(industry_8_9)
print(industry_8_9)


# In[233]:


#2020汽车交通行业存活时间
industry_9 = company_2020[company_2020['Industry'] == '汽车交通']

industry_9_10 = industry_9['Life_time_day']
industry_9_10 = list(industry_9_10)
print(industry_9_10)


# In[234]:


#2020医疗健康行业存活时间
industry_10 = company_2020[company_2020['Industry'] == '医疗健康']

industry_10_11 = industry_10['Life_time_day']
industry_10_11 = list(industry_10_11)
print(industry_10_11)


# In[235]:


#2019企业服务行业存活时间
industry_1 = company_2019[company_2019['Industry'] == '企业服务']
#print(industry_1)

industry_1_3 = industry_1['Life_time_day']
industry_1_3 = list(industry_1_3)
print(industry_1_3)


# In[236]:


#2019电子商务行业存活时间
industry_2 = company_2019[company_2019['Industry'] == '电子商务']

industry_2_4 = industry_2['Life_time_day']
industry_2_4 = list(industry_2_4)
print(industry_2_4)


# In[237]:


#2019教育行业存活时间
industry_3 = company_2019[company_2019['Industry'] == '教育']

industry_3_5 = industry_3['Life_time_day']
industry_3_5 = list(industry_3_5)
print(industry_3_5)


# In[238]:


#2019本地生活行业存活时间
industry_4 = company_2019[company_2019['Industry'] == '本地生活']

industry_4_6 = industry_4['Life_time_day']
industry_4_6 = list(industry_4_6)
print(industry_4_6)


# In[239]:


#2019金融行业存活时间
industry_5 = company_2019[company_2019['Industry'] == '金融']

industry_5_7 = industry_5['Life_time_day']
industry_5_7 = list(industry_5_7)
print(industry_5_7)


# In[240]:


#2019文娱传媒行业存活时间
industry_6 = company_2019[company_2019['Industry'] == '文娱传媒']

industry_6_8 = industry_6['Life_time_day']
industry_6_8 = list(industry_6_8)
print(industry_6_8)


# In[241]:


#2019社交网络行业存活时间
industry_7 = company_2019[company_2019['Industry'] == '社交网络']

industry_7_9 = industry_7['Life_time_day']
industry_7_9 = list(industry_7_9)
print(industry_7_9)


# In[242]:


#2019游戏行业存活时间
industry_8 = company_2019[company_2019['Industry'] == '游戏']

industry_8_10 = industry_8['Life_time_day']
industry_8_10 = list(industry_8_10)
print(industry_8_10)


# In[243]:


#2019汽车交通行业存活时间
industry_9 = company_2019[company_2019['Industry'] == '汽车交通']

industry_9_11 = industry_9['Life_time_day']
industry_9_11 = list(industry_9_11)
print(industry_9_11)


# In[244]:


#2019医疗健康行业存活时间
industry_10 = company_2019[company_2019['Industry'] == '医疗健康']

industry_10_12 = industry_10['Life_time_day']
industry_10_12 = list(industry_10_12)
print(industry_10_12)


# In[246]:


#2019-2021前十行业存活时间合并
industry1 = industry_1_1 + industry_1_2 + industry_1_3
#print(industry1)
industry2 = industry_2_2 + industry_2_3 + industry_2_4
industry3 = industry_3_3 + industry_3_4 + industry_3_5
industry4 = industry_4_4 + industry_4_5 + industry_4_6
industry5 = industry_5_5 + industry_5_6 + industry_5_7
industry6 = industry_6_6 + industry_6_7 + industry_6_8
industry7 = industry_7_7 + industry_7_8 + industry_7_9
industry8 = industry_8_8 + industry_8_9 + industry_8_10
industry9 = industry_9_9 + industry_9_10 + industry_9_11
industry10 = industry_10_10 + industry_10_11 + industry_10_12


# In[249]:


#2019-2021前十行业---存活时间箱型图
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from pylab import *

plt.figure(figsize=(13,13), dpi=200)
rcParams['axes.unicode_minus'] = False
rcParams['font.sans-serif'] = ['Times New Roman']

y1 = pd.Series(np.array(industry1))
y2 = pd.Series(np.array(industry2))
y3 = pd.Series(np.array(industry3))
y4  = pd.Series(np.array(industry4))
y5  = pd.Series(np.array(industry5))
y6  = pd.Series(np.array(industry6))
y7  = pd.Series(np.array(industry7))
y8 = pd.Series(np.array(industry8))
y9 = pd.Series(np.array(industry9))
y10 = pd.Series(np.array(industry10))

data = pd.DataFrame({"Corporation service": y1,
                     "E-commerce ": y2, 
                     "Education": y3, 
                     "Local life": y4,
                     "Fiance": y5,
                     "Entertainment media": y6,
                     "Social network": y7,
                     "Game": y8,
                     "Car transportation": y9,
                     "Medical health": y10,
                    })

data.boxplot()

plt.height = 30
plt.width = 30

#图标题， 设置字号，位置
plt.title('Top 10 industries and Life time (day) of death companies in 2019-2021', fontsize = 20, weight="bold", y = 1.01)
plt.ylabel("Life time (day)", fontsize = 14, weight = "bold")
plt.xlabel("Industry", fontsize = 14, weight = "bold")
plt.savefig("2019-2021industry-lifetime.jpg", dpi = 300, bbox_inches = 'tight')
plt.show()


# In[ ]:




